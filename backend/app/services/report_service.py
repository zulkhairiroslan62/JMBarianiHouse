"""Report Generation Service."""
import os, json
from datetime import datetime, timezone
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.report import Report
from app.models.invoice import Invoice, InvoiceStatus
from app.models.inventory import InventoryItem, StockMovement, MovementType
from app.models.sales import SalesTransaction
from app.schemas.report import ReportRequest
from app.config import settings

def generate_report_file(db: Session, request: ReportRequest, user_id: int) -> Report:
    title = f"{request.report_type.upper()} Report - {request.date_from.strftime('%d/%m/%Y')} to {request.date_to.strftime('%d/%m/%Y')}"
    data = _gather_report_data(db, request)
    summary = _fallback_summary(data, request.report_type)
    filename = f"report_{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    if request.format == "excel":
        filepath = _generate_excel(data, title, summary, filename)
    else:
        filepath = _generate_pdf(data, title, summary, filename)
    report = Report(report_type=request.report_type, title=title, format=request.format, date_from=request.date_from, date_to=request.date_to, filepath=filepath, executive_summary=summary, generated_by=user_id)
    db.add(report)
    db.commit()
    db.refresh(report)
    return report

def _gather_report_data(db: Session, request: ReportRequest) -> dict:
    data = {"type": request.report_type}
    if request.report_type in ["opex", "full"]:
        invoices = db.query(Invoice).filter(Invoice.status.in_([InvoiceStatus.CONFIRMED, InvoiceStatus.PROCESSED]), Invoice.invoice_date >= request.date_from, Invoice.invoice_date <= request.date_to).all()
        data["opex_total"] = sum(i.total_amount or 0 for i in invoices)
        data["invoice_count"] = len(invoices)
    if request.report_type in ["sales", "full"]:
        sales = db.query(SalesTransaction).filter(SalesTransaction.transaction_date >= request.date_from, SalesTransaction.transaction_date <= request.date_to, SalesTransaction.is_void == False).all()
        data["sales_total"] = sum(s.total_price for s in sales)
        data["transaction_count"] = len(sales)
    if request.report_type in ["inventory", "full"]:
        items = db.query(InventoryItem).all()
        data["inventory_items"] = len(items)
        data["below_reorder"] = sum(1 for i in items if i.is_below_reorder)
    if request.report_type in ["waste", "full"]:
        waste = db.query(StockMovement).filter(StockMovement.movement_type == MovementType.WASTE, StockMovement.created_at >= request.date_from, StockMovement.created_at <= request.date_to).all()
        data["waste_total_cost"] = sum(abs(w.total_cost or 0) for w in waste)
    return data

def _fallback_summary(data: dict, report_type: str) -> str:
    parts = []
    if "opex_total" in data:
        parts.append(f"Total OPEX: RM{data['opex_total']:.2f} from {data['invoice_count']} invoices.")
    if "sales_total" in data:
        parts.append(f"Total Sales: RM{data['sales_total']:.2f} ({data['transaction_count']} transactions).")
    if "waste_total_cost" in data:
        parts.append(f"Total Waste: RM{data['waste_total_cost']:.2f}.")
    return " ".join(parts) if parts else "No data for selected period."



def _generate_pdf(data: dict, title: str, summary: str, filename: str) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    filepath = os.path.join(settings.UPLOAD_DIR, f"{filename}.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 12), Paragraph("Executive Summary", styles['Heading2']), Paragraph(summary, styles['Normal']), Spacer(1, 20)]
    if "opex_total" in data:
        elements.append(Paragraph("OPEX Summary", styles['Heading2']))
        table_data = [["Metric", "Value"], ["Total OPEX", f"RM {data['opex_total']:.2f}"], ["Invoice Count", str(data['invoice_count'])]]
        table = Table(table_data)
        table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
        elements.append(table)
    doc.build(elements)
    return filepath

def _generate_excel(data: dict, title: str, summary: str, filename: str) -> str:
    import openpyxl
    from openpyxl.styles import Font
    filepath = os.path.join(settings.UPLOAD_DIR, f"{filename}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A3'] = "Executive Summary:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = summary
    row = 6
    if "opex_total" in data:
        ws.cell(row=row, column=1, value="Total OPEX").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data['opex_total'])
        row += 1
    if "sales_total" in data:
        ws.cell(row=row, column=1, value="Total Sales").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data['sales_total'])
    wb.save(filepath)
    return filepath
